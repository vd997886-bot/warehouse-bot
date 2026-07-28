import asyncio
import difflib
import os
import re
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    Update,
)
from telegram.ext import (
    ApplicationBuilder,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

TOKEN = os.getenv("TOKEN")
ADMIN_ID = os.getenv("ADMIN_ID")
FILE_PATH = Path(os.getenv("WAREHOUSE_FILE", "warehouse.xlsx"))

REQUIRED_COLUMNS = [
    "PartNumber",
    "Quantity",
    "Shelf",
    "Location",
    "Passport",
    "Category",
    "SerialNumber",
    "Check",
]

OPTIONAL_COLUMNS = [
    "Price",
    "PhotoID",
    "SoldTo",
    "SoldDate",
    "Notes",
]

excel_lock = asyncio.Lock()

MAIN_KEYBOARD = ReplyKeyboardMarkup(
    [
        ["🔍 Найти запчасть", "🗑 Удалить запчасть"],
        ["📥 Скачать Excel", "↩️ Отменить удаление"],
        ["❌ Отмена"],
    ],
    resize_keyboard=True,
    is_persistent=True,
)


def authorized(user_id: int) -> bool:
    return not ADMIN_ID or str(user_id) == str(ADMIN_ID)


def safe_str(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def normalize_part(value) -> str:
    text = safe_str(value).upper()
    return re.sub(r"[\s\-_./\\]+", "", text)


def qty_number(value) -> float:
    try:
        return float(safe_str(value).replace(",", "."))
    except Exception:
        return 0.0


def translate(value, field: str) -> str:
    text = safe_str(value).lower()

    if field == "passport":
        if text in {"yes", "y", "true", "1"}:
            return "есть"
        if text in {"no", "n", "false", "0"}:
            return "нет"

    if field == "check":
        if text in {"yes", "y", "true", "1"}:
            return "проверена"
        if text in {"no", "n", "false", "0"}:
            return "не проверена"

    if field == "category":
        return {
            "new": "новая",
            "used": "б/у",
            "serviceable": "исправная",
            "overhauled": "после ремонта",
        }.get(text, safe_str(value))

    return safe_str(value)


def load_dataframe() -> pd.DataFrame:
    if not FILE_PATH.exists():
        raise FileNotFoundError(f"Файл {FILE_PATH.name} не найден")

    df = pd.read_excel(FILE_PATH)
    df.columns = [safe_str(column) for column in df.columns]

    missing = [column for column in REQUIRED_COLUMNS if column not in df.columns]
    if missing:
        raise ValueError("В Excel не хватает обязательных колонок: " + ", ".join(missing))

    for column in OPTIONAL_COLUMNS:
        if column not in df.columns:
            df[column] = ""

    df["PartNumber"] = df["PartNumber"].fillna("").astype(str)
    df["_pn_norm"] = df["PartNumber"].map(normalize_part)
    return df


def workbook_headers(workbook) -> dict[str, int]:
    sheet = workbook.active
    return {
        safe_str(cell.value): index
        for index, cell in enumerate(sheet[1], start=1)
        if safe_str(cell.value)
    }


def create_backup(prefix: str) -> Path:
    if not FILE_PATH.exists():
        raise FileNotFoundError(f"Файл {FILE_PATH.name} не найден")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    backup = FILE_PATH.with_name(f"warehouse_{prefix}_{timestamp}.xlsx")
    shutil.copy2(FILE_PATH, backup)
    return backup


def parse_parts(text: str) -> list[str]:
    text = re.sub(r"^/delete(?:@\w+)?", "", safe_str(text), flags=re.I).strip()
    if not text:
        return []

    result: list[str] = []
    seen: set[str] = set()

    for item in re.split(r"[,;\n]+", text):
        item = item.strip()
        normalized = normalize_part(item)
        if item and normalized and normalized not in seen:
            seen.add(normalized)
            result.append(item)

    return result


def find_rows(parts: list[str]) -> tuple[list[dict], list[str]]:
    workbook = load_workbook(FILE_PATH, data_only=False)
    sheet = workbook.active
    headers = workbook_headers(workbook)

    part_col = headers.get("PartNumber")
    if not part_col:
        workbook.close()
        raise ValueError("В Excel нет колонки PartNumber")

    requested = {normalize_part(part): part for part in parts}
    found_norm: set[str] = set()
    rows: list[dict] = []

    def value(row: int, column_name: str):
        column = headers.get(column_name)
        return sheet.cell(row=row, column=column).value if column else ""

    for row_number in range(2, sheet.max_row + 1):
        part_value = value(row_number, "PartNumber")
        normalized = normalize_part(part_value)
        if normalized not in requested:
            continue

        found_norm.add(normalized)
        rows.append(
            {
                "row": row_number,
                "part_number": safe_str(part_value),
                "normalized": normalized,
                "serial_number": safe_str(value(row_number, "SerialNumber")),
                "quantity": safe_str(value(row_number, "Quantity")),
                "shelf": safe_str(value(row_number, "Shelf")),
                "location": safe_str(value(row_number, "Location")),
            }
        )

    workbook.close()
    not_found = [original for normalized, original in requested.items() if normalized not in found_norm]
    return rows, not_found


def verify_rows(rows: list[dict]) -> bool:
    workbook = load_workbook(FILE_PATH, data_only=False)
    sheet = workbook.active
    headers = workbook_headers(workbook)
    part_col = headers.get("PartNumber")

    if not part_col:
        workbook.close()
        return False

    for item in rows:
        row_number = item["row"]
        if row_number < 2 or row_number > sheet.max_row:
            workbook.close()
            return False

        current = sheet.cell(row=row_number, column=part_col).value
        if normalize_part(current) != item["normalized"]:
            workbook.close()
            return False

    workbook.close()
    return True


def delete_rows(rows: list[dict]) -> list[dict]:
    workbook = load_workbook(FILE_PATH)
    sheet = workbook.active
    deleted: list[dict] = []

    for item in sorted(rows, key=lambda row: row["row"], reverse=True):
        row_number = item["row"]
        if 2 <= row_number <= sheet.max_row:
            sheet.delete_rows(row_number, 1)
            deleted.append(item)

    workbook.save(FILE_PATH)
    workbook.close()
    return sorted(deleted, key=lambda row: row["row"])


def format_row(row) -> str:
    part = safe_str(row.get("PartNumber"))
    quantity = safe_str(row.get("Quantity"))
    shelf = safe_str(row.get("Shelf"))
    location = safe_str(row.get("Location"))
    passport = translate(row.get("Passport"), "passport")
    category = translate(row.get("Category"), "category")
    serial = safe_str(row.get("SerialNumber")) or "—"
    check = translate(row.get("Check"), "check")
    price = safe_str(row.get("Price")) or "—"

    status = "❌ ПРОДАНО" if qty_number(row.get("Quantity")) <= 0 else f"✅ {part} есть в наличии"
    if status == "❌ ПРОДАНО":
        status = f"❌ ПРОДАНО\n📦 {part}"

    text = (
        f"{status}\n"
        f"📍 Полка: {shelf}, ячейка: {location}\n"
        f"🔢 Количество: {quantity}\n"
        f"📄 Паспорт: {passport}\n"
        f"🆕 Категория: {category}\n"
        f"💰 Цена: {price}\n"
        f"🔑 Серийный номер: {serial}\n"
        f"✔ Проверка: {check}"
    )

    sold_to = safe_str(row.get("SoldTo"))
    sold_date = safe_str(row.get("SoldDate"))
    notes = safe_str(row.get("Notes"))

    if sold_to:
        text += f"\n👤 Кому продано: {sold_to}"
    if sold_date:
        text += f"\n📅 Дата продажи: {sold_date.split(' ')[0]}"
    if notes:
        text += f"\n📝 Заметка: {notes}"

    return text


def delete_preview(found: list[dict], not_found: list[str]) -> str:
    lines = ["⚠️ Будут полностью удалены:", ""]

    for item in found[:40]:
        line = f"📦 {item['part_number']}"
        if item["serial_number"]:
            line += f" | S/N: {item['serial_number']}"
        if item["quantity"]:
            line += f" | Qty: {item['quantity']}"
        lines.append(line)

    if len(found) > 40:
        lines += ["", f"…и ещё {len(found) - 40} строк."]

    if not_found:
        lines += ["", "❌ Не найдены:"]
        lines += [f"• {part}" for part in not_found[:30]]

    lines += ["", f"Всего строк будет удалено: {len(found)}"]
    return "\n".join(lines)


async def send_excel(context: ContextTypes.DEFAULT_TYPE, chat_id: int, caption: str) -> None:
    if not FILE_PATH.exists():
        await context.bot.send_message(chat_id=chat_id, text="❌ Excel-файл не найден")
        return

    with FILE_PATH.open("rb") as file:
        await context.bot.send_document(
            chat_id=chat_id,
            document=file,
            filename="warehouse_actual.xlsx",
            caption=caption,
            reply_markup=MAIN_KEYBOARD,
        )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data.clear()
    await update.message.reply_text(
        "Привет! 👋 Выбери действие кнопкой внизу.",
        reply_markup=MAIN_KEYBOARD,
    )


async def begin_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data["mode"] = "search"
    await update.message.reply_text(
        "🔍 Отправь PartNumber или часть номера.",
        reply_markup=MAIN_KEYBOARD,
    )


async def begin_delete(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Нет доступа", reply_markup=MAIN_KEYBOARD)
        return

    context.user_data["mode"] = "delete"
    await update.message.reply_text(
        "🗑 Отправь PartNumber. Несколько номеров можно через запятую или с новой строки.",
        reply_markup=MAIN_KEYBOARD,
    )


async def search_part(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str) -> None:
    query = normalize_part(text)
    if not query:
        await update.message.reply_text("❓ Напиши PartNumber", reply_markup=MAIN_KEYBOARD)
        return

    try:
        df = load_dataframe()
    except Exception as error:
        await update.message.reply_text(f"⚠️ Ошибка: {error}", reply_markup=MAIN_KEYBOARD)
        return

    exact = df[df["_pn_norm"] == query]
    matches = exact

    if matches.empty:
        matches = df[df["_pn_norm"].str.contains(query, na=False, regex=False)]

    if matches.empty:
        close = difflib.get_close_matches(query, df["_pn_norm"].tolist(), n=10, cutoff=0.75)
        if close:
            matches = df[df["_pn_norm"].isin(close)]

    if matches.empty:
        await update.message.reply_text("❓ Ничего не найдено", reply_markup=MAIN_KEYBOARD)
        return

    for _, row in matches.head(10).iterrows():
        photo_id = safe_str(row.get("PhotoID"))
        caption = format_row(row)

        if photo_id:
            try:
                await update.message.reply_photo(photo=photo_id, caption=caption, reply_markup=MAIN_KEYBOARD)
                continue
            except Exception:
                pass

        await update.message.reply_text(caption, reply_markup=MAIN_KEYBOARD)

    if len(matches) > 10:
        await update.message.reply_text(
            f"ℹ️ Всего найдено {len(matches)}. Показаны первые 10.",
            reply_markup=MAIN_KEYBOARD,
        )


async def prepare_delete(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str) -> None:
    parts = parse_parts(text)
    if not parts:
        await update.message.reply_text("❌ Не удалось прочитать PartNumber", reply_markup=MAIN_KEYBOARD)
        return

    try:
        async with excel_lock:
            found, not_found = find_rows(parts)
    except Exception as error:
        await update.message.reply_text(f"⚠️ Ошибка: {error}", reply_markup=MAIN_KEYBOARD)
        return

    if not found:
        await update.message.reply_text("❌ Ничего не найдено для удаления", reply_markup=MAIN_KEYBOARD)
        return

    context.user_data["pending_delete"] = found

    keyboard = InlineKeyboardMarkup(
        [[
            InlineKeyboardButton("✅ Удалить", callback_data="confirm_delete"),
            InlineKeyboardButton("❌ Отмена", callback_data="cancel_delete"),
        ]]
    )
    await update.message.reply_text(delete_preview(found, not_found), reply_markup=keyboard)


async def delete_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    if not authorized(query.from_user.id):
        await query.edit_message_text("⛔ Нет доступа")
        return

    if query.data == "cancel_delete":
        context.user_data.pop("pending_delete", None)
        context.user_data["mode"] = None
        await query.edit_message_text("❌ Удаление отменено")
        return

    pending = context.user_data.get("pending_delete")
    if not pending:
        await query.edit_message_text("⚠️ Список удаления устарел. Начни заново.")
        return

    try:
        async with excel_lock:
            if not verify_rows(pending):
                raise RuntimeError("Excel изменился. Начни удаление заново.")

            backup = create_backup("before_delete")
            deleted = delete_rows(pending)
            context.application.bot_data["last_backup"] = str(backup)
    except Exception as error:
        await query.edit_message_text(f"⚠️ Ошибка удаления: {error}")
        return

    context.user_data.clear()
    await query.edit_message_text(f"✅ Удалено строк: {len(deleted)}")
    await send_excel(
        context,
        query.message.chat_id,
        "📥 Обновлённый Excel. Удалённые позиции уже убраны.",
    )


async def undo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Нет доступа", reply_markup=MAIN_KEYBOARD)
        return

    backup = context.application.bot_data.get("last_backup")
    if not backup or not Path(backup).exists():
        await update.message.reply_text("❌ Нечего отменять", reply_markup=MAIN_KEYBOARD)
        return

    try:
        async with excel_lock:
            create_backup("before_undo")
            shutil.copy2(backup, FILE_PATH)
            context.application.bot_data.pop("last_backup", None)
    except Exception as error:
        await update.message.reply_text(f"⚠️ Ошибка восстановления: {error}", reply_markup=MAIN_KEYBOARD)
        return

    await update.message.reply_text("✅ Последнее удаление отменено", reply_markup=MAIN_KEYBOARD)
    await send_excel(context, update.effective_chat.id, "📥 Восстановленный Excel")


async def download_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Нет доступа", reply_markup=MAIN_KEYBOARD)
        return

    async with excel_lock:
        await send_excel(context, update.effective_chat.id, "📥 Актуальный склад")


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not authorized(update.effective_user.id):
        await update.message.reply_text("⛔ Нет доступа", reply_markup=MAIN_KEYBOARD)
        return

    document = update.message.document
    if not document or not (document.file_name or "").lower().endswith(".xlsx"):
        await update.message.reply_text("❌ Пришли файл .xlsx", reply_markup=MAIN_KEYBOARD)
        return

    temp = FILE_PATH.with_name("warehouse_upload_temp.xlsx")

    try:
        telegram_file = await context.bot.get_file(document.file_id)
        await telegram_file.download_to_drive(temp)

        test_df = pd.read_excel(temp)
        test_df.columns = [safe_str(column) for column in test_df.columns]
        missing = [column for column in REQUIRED_COLUMNS if column not in test_df.columns]
        if missing:
            raise ValueError("Не хватает колонок: " + ", ".join(missing))

        async with excel_lock:
            if FILE_PATH.exists():
                create_backup("before_upload")
            shutil.move(temp, FILE_PATH)
            load_dataframe()
    except Exception as error:
        temp.unlink(missing_ok=True)
        await update.message.reply_text(f"⚠️ Не удалось загрузить Excel: {error}", reply_markup=MAIN_KEYBOARD)
        return

    context.user_data.clear()
    await update.message.reply_text("✅ Таблица обновлена", reply_markup=MAIN_KEYBOARD)


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    photo_id = update.message.photo[-1].file_id
    await update.message.reply_text(
        f"PhotoID:\n{photo_id}",
        reply_markup=MAIN_KEYBOARD,
    )


async def delete_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = re.sub(r"^/delete(?:@\w+)?", "", update.message.text or "", flags=re.I).strip()
    if text:
        await prepare_delete(update, context, text)
    else:
        await begin_delete(update, context)


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = safe_str(update.message.text)

    if text == "🔍 Найти запчасть":
        await begin_search(update, context)
    elif text == "🗑 Удалить запчасть":
        await begin_delete(update, context)
    elif text == "📥 Скачать Excel":
        context.user_data["mode"] = None
        await download_excel(update, context)
    elif text == "↩️ Отменить удаление":
        await undo(update, context)
    elif text == "❌ Отмена":
        context.user_data.clear()
        await update.message.reply_text("❌ Действие отменено", reply_markup=MAIN_KEYBOARD)
    elif context.user_data.get("mode") == "delete":
        await prepare_delete(update, context, text)
    else:
        await search_part(update, context, text)


def main() -> None:
    if not TOKEN:
        raise RuntimeError("Добавь TOKEN в Railway Variables")

    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", start))
    app.add_handler(CommandHandler("delete", delete_command))
    app.add_handler(CommandHandler("undo", undo))
    app.add_handler(CallbackQueryHandler(delete_callback, pattern=r"^(confirm_delete|cancel_delete)$"))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    print("🤖 Warehouse bot started")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
    "PartNumber",
    "Quantity",
    "Shelf",
    "Location",
    "Passport",
    "Category",
    "SerialNumber",
    "Check",
]

# Эти колонки могут отсутствовать.
# Бот автоматически будет считать их пустыми.
OPTIONAL_COLUMNS = [
    "Price",
    "PhotoID",
    "SoldTo",
    "SoldDate",
    "Notes",
]

excel_lock = asyncio.Lock()


# =========================================================
# ГЛАВНОЕ МЕНЮ
# =========================================================

MAIN_KEYBOARD = ReplyKeyboardMarkup(
    [
        [
            "🔍 Найти запчасть",
            "🗑 Удалить запчасть",
        ],
        [
            "📥 Скачать Excel",
            "↩️ Отменить удаление",
        ],
        [
            "❌ Отмена",
        ],
    ],
    resize_keyboard=True,
    is_persistent=True,
)


# =========================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =========================================================

def is_authorized(user_id: int) -> bool:
    """
    Если ADMIN_ID не добавлен в Railway,
    доступ к управлению открыт всем.

    Если ADMIN_ID добавлен,
    управлять складом может только этот пользователь.
    """

    if not ADMIN_ID:
        return True

    return str(user_id) == str(ADMIN_ID)


def normalize_part_for_search(value) -> str:
    """
    Приводит PartNumber к единому виду.

    Например:
    ПУ-11
    ПУ 11
    ПУ_11

    будут считаться одинаковыми.
    """

    if value is None:
        return ""

    value = str(value).strip().upper()

    value = re.sub(
        r"[\s\-_./\\]+",
        "",
        value,
    )

    return value


def safe_str(value) -> str:
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    return str(value).strip()


def translate_value(value, field: str) -> str:
    value_text = safe_str(value).lower()

    if field == "passport":
        if value_text in ["yes", "y", "true", "1"]:
            return "есть"

        if value_text in ["no", "n", "false", "0"]:
            return "нет"

    if field == "check":
        if value_text in ["yes", "y", "true", "1"]:
            return "проверена"

        if value_text in ["no", "n", "false", "0"]:
            return "не проверена"

    if field == "category":
        if value_text == "new":
            return "новая"

        if value_text == "used":
            return "б/у"

        if value_text == "serviceable":
            return "исправная"

        if value_text == "overhauled":
            return "после ремонта"

    return safe_str(value)


def clean_serial(value) -> str:
    serial = safe_str(value)

    if not serial or serial in ["/", "-", "—"]:
        return "—"

    return serial


def clean_price(value) -> str:
    price = safe_str(value)

    if not price or price in ["/", "-", "—"]:
        return "—"

    price = price.replace("USD", "$")
    price = price.replace("usd", "$")

    return price.strip()


def clean_date(value) -> str:
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    try:
        date_value = pd.to_datetime(value)

        return date_value.strftime("%Y-%m-%d")

    except Exception:
        return str(value).split(" ")[0]


def qty_to_number(value) -> float:
    quantity = safe_str(value).replace(",", ".")

    try:
        return float(quantity)

    except Exception:
        return 0.0


# =========================================================
# ЧТЕНИЕ EXCEL
# =========================================================

def load_df() -> pd.DataFrame:
    if not os.path.exists(FILE_PATH):
        raise FileNotFoundError(
            f"Файл {FILE_PATH} не найден. "
            "Отправь Excel .xlsx файлом в бота."
        )

    dataframe = pd.read_excel(FILE_PATH)

    dataframe.columns = [
        str(column).strip()
        for column in dataframe.columns
    ]

    missing_columns = [
        column
        for column in REQUIRED_COLUMNS
        if column not in dataframe.columns
    ]

    if missing_columns:
        raise ValueError(
            "В Excel не хватает обязательных колонок:\n"
            + ", ".join(missing_columns)
        )

    # Необязательные колонки создаются только внутри программы.
    # Сам Excel из-за этого не перезаписывается.
    for column in OPTIONAL_COLUMNS:
        if column not in dataframe.columns:
            dataframe[column] = ""

    dataframe["PartNumber"] = (
        dataframe["PartNumber"]
        .fillna("")
        .astype(str)
    )

    dataframe["_pn_norm"] = dataframe[
        "PartNumber"
    ].apply(
        normalize_part_for_search
    )

    return dataframe


def get_excel_headers(workbook) -> dict:
    worksheet = workbook.active
    headers = {}

    for column_number, cell in enumerate(
        worksheet[1],
        start=1,
    ):
        if cell.value is not None:
            header_name = str(cell.value).strip()
            headers[header_name] = column_number

    return headers


# =========================================================
# РЕЗЕРВНЫЕ КОПИИ
# =========================================================

def create_backup(prefix: str = "backup") -> str:
    if not os.path.exists(FILE_PATH):
        raise FileNotFoundError(
            f"Файл {FILE_PATH} не найден."
        )

    timestamp = datetime.now().strftime(
        "%Y-%m-%d_%H-%M-%S_%f"
    )

    backup_path = (
        f"warehouse_{prefix}_{timestamp}.xlsx"
    )

    shutil.copy2(
        FILE_PATH,
        backup_path,
    )

    return backup_path


# =========================================================
# УДАЛЕНИЕ ПО PARTNUMBER
# =========================================================

def parse_part_numbers(text: str) -> list[str]:
    """
    Поддерживает:

    ПУ-11

    ПУ-11, ЭЦН-333М, БПСР4

    Или каждый PartNumber с новой строки.
    """

    text = safe_str(text)

    text = re.sub(
        r"^/delete(?:@\w+)?",
        "",
        text,
        flags=re.IGNORECASE,
    ).strip()

    if not text:
        return []

    parts = re.split(
        r"[,;\n]+",
        text,
    )

    result = []
    seen = set()

    for part in parts:
        part = part.strip()

        if not part:
            continue

        normalized = normalize_part_for_search(
            part
        )

        if normalized and normalized not in seen:
            seen.add(normalized)
            result.append(part)

    return result


def find_rows_by_part_numbers(
    requested_parts: list[str],
) -> tuple[list[dict], list[str]]:
    """
    Ищет точное совпадение PartNumber.

    Если один PartNumber встречается несколько раз,
    бот найдёт все строки с этим номером.
    """

    if not os.path.exists(FILE_PATH):
        raise FileNotFoundError(
            f"Файл {FILE_PATH} не найден."
        )

    workbook = load_workbook(
        FILE_PATH,
        data_only=False,
    )

    worksheet = workbook.active
    headers = get_excel_headers(workbook)

    part_column = headers.get("PartNumber")
    serial_column = headers.get("SerialNumber")
    quantity_column = headers.get("Quantity")
    shelf_column = headers.get("Shelf")
    location_column = headers.get("Location")

    if not part_column:
        workbook.close()

        raise ValueError(
            "В Excel не найдена колонка PartNumber."
        )

    requested_map = {
        normalize_part_for_search(part): part
        for part in requested_parts
    }

    found_normalized = set()
    found_rows = []

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        part_value = worksheet.cell(
            row=row_number,
            column=part_column,
        ).value

        normalized = normalize_part_for_search(
            part_value
        )

        if normalized not in requested_map:
            continue

        found_normalized.add(normalized)

        serial_value = ""

        if serial_column:
            serial_value = worksheet.cell(
                row=row_number,
                column=serial_column,
            ).value

        quantity_value = ""

        if quantity_column:
            quantity_value = worksheet.cell(
                row=row_number,
                column=quantity_column,
            ).value

        shelf_value = ""

        if shelf_column:
            shelf_value = worksheet.cell(
                row=row_number,
                column=shelf_column,
            ).value

        location_value = ""

        if location_column:
            location_value = worksheet.cell(
                row=row_number,
                column=location_column,
            ).value

        found_rows.append(
            {
                "row": row_number,
                "part_number": safe_str(part_value),
                "normalized": normalized,
                "serial_number": safe_str(
                    serial_value
                ),
                "quantity": safe_str(
                    quantity_value
                ),
                "shelf": safe_str(
                    shelf_value
                ),
                "location": safe_str(
                    location_value
                ),
            }
        )

    workbook.close()

    not_found = []

    for normalized, original in requested_map.items():
        if normalized not in found_normalized:
            not_found.append(original)

    return found_rows, not_found


def verify_pending_rows(
    pending_rows: list[dict],
) -> bool:
    """
    Проверяет, не изменился ли Excel
    перед подтверждением удаления.
    """

    if not os.path.exists(FILE_PATH):
        return False

    workbook = load_workbook(
        FILE_PATH,
        data_only=False,
    )

    worksheet = workbook.active
    headers = get_excel_headers(workbook)

    part_column = headers.get("PartNumber")

    if not part_column:
        workbook.close()
        return False

    for item in pending_rows:
        row_number = item["row"]

        if (
            row_number < 2
            or row_number > worksheet.max_row
        ):
            workbook.close()
            return False

        current_part = worksheet.cell(
            row=row_number,
            column=part_column,
        ).value

        current_normalized = (
            normalize_part_for_search(
                current_part
            )
        )

        if (
            current_normalized
            != item["normalized"]
        ):
            workbook.close()
            return False

    workbook.close()

    return True


def delete_excel_rows(
    rows_to_delete: list[dict],
) -> list[dict]:
    """
    Полностью удаляет строки из Excel.

    Удаляет снизу вверх,
    чтобы номера строк не смещались.
    """

    if not os.path.exists(FILE_PATH):
        raise FileNotFoundError(
            f"Файл {FILE_PATH} не найден."
        )

    workbook = load_workbook(FILE_PATH)
    worksheet = workbook.active

    deleted_items = []

    sorted_rows = sorted(
        rows_to_delete,
        key=lambda item: item["row"],
        reverse=True,
    )

    for item in sorted_rows:
        row_number = item["row"]

        if (
            row_number < 2
            or row_number > worksheet.max_row
        ):
            continue

        deleted_items.append(item)

        worksheet.delete_rows(
            row_number,
            1,
        )

    workbook.save(FILE_PATH)
    workbook.close()

    return sorted(
        deleted_items,
        key=lambda item: item["row"],
    )


def format_delete_preview(
    found_rows: list[dict],
    not_found: list[str],
) -> str:
    lines = [
        "⚠️ Будут полностью удалены:",
        "",
    ]

    for item in found_rows[:40]:
        line = f'📦 {item["part_number"]}'

        if item["serial_number"]:
            line += (
                f' | S/N: {item["serial_number"]}'
            )

        if item["quantity"]:
            line += (
                f' | Qty: {item["quantity"]}'
            )

        if item["shelf"]:
            line += (
                f' | Полка: {item["shelf"]}'
            )

        if item["location"]:
            line += (
                f' | Ячейка: {item["location"]}'
            )

        lines.append(line)

    if len(found_rows) > 40:
        lines.append("")
        lines.append(
            f"…и ещё {len(found_rows) - 40} строк."
        )

    if not_found:
        lines.append("")
        lines.append("❌ Не найдены:")

        for part in not_found[:30]:
            lines.append(f"• {part}")

        if len(not_found) > 30:
            lines.append(
                f"…и ещё {len(not_found) - 30}."
            )

    lines.append("")
    lines.append(
        f"Всего будет удалено строк: "
        f"{len(found_rows)}"
    )

    lines.append("")
    lines.append(
        "После удаления запчасти исчезнут "
        "из Excel и из поиска бота."
    )

    return "\n".join(lines)


# =========================================================
# ФОРМАТ ОТВЕТА ПРИ ПОИСКЕ
# =========================================================

def fmt_row(row) -> str:
    part = safe_str(
        row.get("PartNumber")
    )

    quantity = safe_str(
        row.get("Quantity")
    )

    shelf = safe_str(
        row.get("Shelf")
    )

    location = safe_str(
        row.get("Location")
    )

    passport = translate_value(
        row.get("Passport"),
        "passport",
    )

    category = translate_value(
        row.get("Category"),
        "category",
    )

    check = translate_value(
        row.get("Check"),
        "check",
    )

    serial = clean_serial(
        row.get("SerialNumber")
    )

    price = clean_price(
        row.get("Price")
    )

    sold_to = safe_str(
        row.get("SoldTo")
    )

    sold_date = clean_date(
        row.get("SoldDate")
    )

    notes = safe_str(
        row.get("Notes")
    )

    quantity_number = qty_to_number(
        row.get("Quantity")
    )

    if quantity_number <= 0:
        text = (
            f"❌ ПРОДАНО\n"
            f"📦 {part}\n"
            f"📍 Полка: {shelf}, "
            f"ячейка: {location}\n"
            f"🔢 Количество: {quantity}\n"
            f"📄 Паспорт: {passport}\n"
            f"🆕 Категория: {category}\n"
            f"💰 Цена: {price}\n"
            f"🔑 Серийный номер: {serial}\n"
            f"✔ Проверка: {check}"
        )

        if sold_to:
            text += (
                f"\n👤 Кому продано: {sold_to}"
            )

        if sold_date:
            text += (
                f"\n📅 Дата продажи: {sold_date}"
            )

        if notes:
            text += (
                f"\n📝 Заметка: {notes}"
            )

        return text

    return (
        f"✅ {part} есть в наличии\n"
        f"📦 Полка: {shelf}, "
        f"ячейка: {location}\n"
        f"🔢 Количество: {quantity}\n"
        f"📄 Паспорт: {passport}\n"
        f"🆕 Категория: {category}\n"
        f"💰 Цена: {price}\n"
        f"🔑 Серийный номер: {serial}\n"
        f"✔ Проверка: {check}"
    )


async def send_part_response(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    row,
):
    caption = fmt_row(row)

    photo_id = safe_str(
        row.get("PhotoID")
    )

    if photo_id:
        try:
            await update.message.reply_photo(
                photo=photo_id,
                caption=caption,
                reply_markup=MAIN_KEYBOARD,
            )
            return

        except Exception as error:
            print(
                "PHOTO ERROR:",
                error,
            )

    await update.message.reply_text(
        caption,
        reply_markup=MAIN_KEYBOARD,
    )


# =========================================================
# ОТПРАВКА АКТУАЛЬНОГО EXCEL
# =========================================================

async def send_current_excel_to_chat(
    context: ContextTypes.DEFAULT_TYPE,
    chat_id: int,
    caption: str,
):
    if not os.path.exists(FILE_PATH):
        await context.bot.send_message(
            chat_id=chat_id,
            text=(
                "❌ Файл warehouse.xlsx "
                "не найден."
            ),
            reply_markup=MAIN_KEYBOARD,
        )
        return

    with open(FILE_PATH, "rb") as excel_file:
        await context.bot.send_document(
            chat_id=chat_id,
            document=excel_file,
            filename="warehouse_actual.xlsx",
            caption=caption,
            reply_markup=MAIN_KEYBOARD,
        )


async def download_excel(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    user = update.effective_user

    if (
        not user
        or not is_authorized(user.id)
    ):
        await update.message.reply_text(
            "⛔ У вас нет доступа "
            "к скачиванию базы.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    try:
        async with excel_lock:
            await send_current_excel_to_chat(
                context=context,
                chat_id=update.effective_chat.id,
                caption=(
                    "📥 Актуальный склад.\n\n"
                    "В этом файле уже нет "
                    "удалённых запчастей."
                ),
            )

    except Exception as error:
        await update.message.reply_text(
            "⚠️ Не удалось отправить Excel:\n"
            f"{error}",
            reply_markup=MAIN_KEYBOARD,
        )


# =========================================================
# КОМАНДЫ И КНОПКИ
# =========================================================

async def start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    context.user_data["mode"] = None

    context.user_data.pop(
        "pending_delete_rows",
        None,
    )

    await update.message.reply_text(
        "Привет! 👋\n\n"
        "Выбери нужное действие кнопкой внизу.",
        reply_markup=MAIN_KEYBOARD,
    )


async def help_cmd(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    await update.message.reply_text(
        "🔍 Найти запчасть — поиск по PartNumber.\n\n"
        "🗑 Удалить запчасть — удалить "
        "одну или несколько позиций.\n\n"
        "📥 Скачать Excel — получить "
        "актуальный список склада.\n\n"
        "↩️ Отменить удаление — восстановить "
        "последнее удаление.\n\n"
        "Чтобы загрузить новый склад, просто "
        "отправь боту Excel-файл .xlsx.",
        reply_markup=MAIN_KEYBOARD,
    )


async def begin_search_mode(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    context.user_data["mode"] = "search"

    await update.message.reply_text(
        "🔍 Отправь PartNumber "
        "или часть номера.\n\n"
        "Например:\n"
        "ПУ-11",
        reply_markup=MAIN_KEYBOARD,
    )


async def begin_delete_mode(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    user = update.effective_user

    if (
        not user
        or not is_authorized(user.id)
    ):
        await update.message.reply_text(
            "⛔ У вас нет доступа к удалению.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    context.user_data["mode"] = "delete"

    await update.message.reply_text(
        "🗑 Отправь PartNumber, "
        "который нужно удалить.\n\n"
        "Одна позиция:\n"
        "ПУ-11\n\n"
        "Несколько позиций:\n"
        "ПУ-11, ЭЦН-333М, БПСР4",
        reply_markup=MAIN_KEYBOARD,
    )


async def cancel_mode(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    context.user_data["mode"] = None

    context.user_data.pop(
        "pending_delete_rows",
        None,
    )

    await update.message.reply_text(
        "❌ Действие отменено.",
        reply_markup=MAIN_KEYBOARD,
    )


# =========================================================
# ПОДГОТОВКА УДАЛЕНИЯ
# =========================================================

async def prepare_delete(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    text: str,
):
    requested_parts = parse_part_numbers(
        text
    )

    if not requested_parts:
        await update.message.reply_text(
            "❌ Не удалось прочитать PartNumber.\n\n"
            "Например:\n"
            "ПУ-11, ЭЦН-333М",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    if len(requested_parts) > 500:
        await update.message.reply_text(
            "❌ За один раз можно указать "
            "максимум 500 PartNumber.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    try:
        async with excel_lock:
            found_rows, not_found = (
                find_rows_by_part_numbers(
                    requested_parts
                )
            )

    except Exception as error:
        await update.message.reply_text(
            "⚠️ Ошибка при чтении Excel:\n"
            f"{error}",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    if not found_rows:
        message = (
            "❌ Ни одна запчасть не найдена.\n\n"
            "Проверь написание PartNumber."
        )

        if not_found:
            message += "\n\nНе найдены:\n"

            message += "\n".join(
                f"• {part}"
                for part in not_found[:30]
            )

        await update.message.reply_text(
            message,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    context.user_data[
        "pending_delete_rows"
    ] = found_rows

    keyboard = InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "✅ Удалить",
                    callback_data="confirm_delete",
                ),
                InlineKeyboardButton(
                    "❌ Отмена",
                    callback_data="cancel_delete",
                ),
            ]
        ]
    )

    await update.message.reply_text(
        format_delete_preview(
            found_rows,
            not_found,
        ),
        reply_markup=keyboard,
    )


async def delete_cmd(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    user = update.effective_user

    if (
        not user
        or not is_authorized(user.id)
    ):
        await update.message.reply_text(
            "⛔ У вас нет доступа к удалению.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    command_text = update.message.text or ""

    part_text = re.sub(
        r"^/delete(?:@\w+)?",
        "",
        command_text,
        flags=re.IGNORECASE,
    ).strip()

    if not part_text:
        await begin_delete_mode(
            update,
            context,
        )
        return

    await prepare_delete(
        update,
        context,
        part_text,
    )


# =========================================================
# ПОДТВЕРЖДЕНИЕ УДАЛЕНИЯ
# =========================================================

async def delete_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    query = update.callback_query

    if not query:
        return

    await query.answer()

    user = query.from_user

    if not is_authorized(user.id):
        await query.edit_message_text(
            "⛔ У вас нет доступа к удалению."
        )
        return

    if query.data == "cancel_delete":
        context.user_data["mode"] = None

        context.user_data.pop(
            "pending_delete_rows",
            None,
        )

        await query.edit_message_text(
            "❌ Удаление отменено. "
            "Excel не изменён."
        )
        return

    if query.data != "confirm_delete":
        return

    pending_rows = context.user_data.get(
        "pending_delete_rows"
    )

    if not pending_rows:
        await query.edit_message_text(
            "⚠️ Список удаления уже недействителен.\n"
            "Нажми кнопку «🗑 Удалить запчасть» "
            "ещё раз."
        )
        return

    try:
        async with excel_lock:
            if not verify_pending_rows(
                pending_rows
            ):
                context.user_data.pop(
                    "pending_delete_rows",
                    None,
                )

                context.user_data["mode"] = None

                await query.edit_message_text(
                    "⚠️ Excel изменился после поиска.\n"
                    "Для безопасности удаление "
                    "остановлено.\n\n"
                    "Попробуй ещё раз."
                )
                return

            backup_path = create_backup(
                "before_delete"
            )

            deleted_items = delete_excel_rows(
                pending_rows
            )

            context.application.bot_data[
                "last_backup"
            ] = backup_path

            context.application.bot_data[
                "last_delete_user"
            ] = user.id

    except Exception as error:
        await query.edit_message_text(
            "⚠️ Ошибка при удалении:\n"
            f"{error}"
        )
        return

    context.user_data.pop(
        "pending_delete_rows",
        None,
    )

    context.user_data["mode"] = None

    lines = [
        "✅ Полностью удалено строк: "
        f"{len(deleted_items)}",
        "",
    ]

    for item in deleted_items[:40]:
        line = f'• {item["part_number"]}'

        if item["serial_number"]:
            line += (
                f' | S/N: {item["serial_number"]}'
            )

        lines.append(line)

    if len(deleted_items) > 40:
        lines.append("")
        lines.append(
            f"…и ещё {len(deleted_items) - 40}."
        )

    lines.append("")
    lines.append(
        "Эти запчасти больше не находятся "
        "через поиск."
    )

    lines.append("")
    lines.append(
        "Сейчас отправляю новый актуальный Excel."
    )

    await query.edit_message_text(
        "\n".join(lines)
    )

    try:
        await send_current_excel_to_chat(
            context=context,
            chat_id=query.message.chat_id,
            caption=(
                "📥 Обновлённый Excel.\n\n"
                "Удалённые запчасти уже убраны "
                "из списка.\n"
                "Этот файл можно отправлять клиентам."
            ),
        )

    except Exception as error:
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=(
                "⚠️ Запчасти удалены, но не удалось "
                "отправить Excel:\n"
                f"{error}\n\n"
                "Нажми кнопку «📥 Скачать Excel»."
            ),
            reply_markup=MAIN_KEYBOARD,
        )


# =========================================================
# ОТМЕНА ПОСЛЕДНЕГО УДАЛЕНИЯ
# =========================================================

async def undo_cmd(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    user = update.effective_user

    if (
        not user
        or not is_authorized(user.id)
    ):
        await update.message.reply_text(
            "⛔ У вас нет доступа "
            "к восстановлению.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    backup_path = (
        context.application.bot_data.get(
            "last_backup"
        )
    )

    if (
        not backup_path
        or not os.path.exists(backup_path)
    ):
        await update.message.reply_text(
            "❌ Нет последнего удаления, "
            "которое можно отменить.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    try:
        async with excel_lock:
            create_backup(
                "before_undo"
            )

            shutil.copy2(
                backup_path,
                FILE_PATH,
            )

            context.application.bot_data.pop(
                "last_backup",
                None,
            )

            context.application.bot_data.pop(
                "last_delete_user",
                None,
            )

    except Exception as error:
        await update.message.reply_text(
            "⚠️ Не удалось восстановить Excel:\n"
            f"{error}",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    context.user_data["mode"] = None

    await update.message.reply_text(
        "✅ Последнее удаление отменено.\n"
        "Предыдущая версия Excel восстановлена.",
        reply_markup=MAIN_KEYBOARD,
    )

    try:
        await send_current_excel_to_chat(
            context=context,
            chat_id=update.effective_chat.id,
            caption=(
                "📥 Восстановленный Excel.\n\n"
                "Удалённые позиции снова находятся "
                "в списке."
            ),
        )

    except Exception as error:
        await update.message.reply_text(
            "⚠️ Excel восстановлен, "
            "но не удалось отправить файл:\n"
            f"{error}",
            reply_markup=MAIN_KEYBOARD,
        )


# =========================================================
# ЗАГРУЗКА НОВОГО EXCEL
# =========================================================

async def handle_document(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    user = update.effective_user

    if (
        not user
        or not is_authorized(user.id)
    ):
        await update.message.reply_text(
            "⛔ У вас нет доступа "
            "к обновлению базы.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    document = update.message.document

    if not document:
        return

    file_name = document.file_name or ""

    if not file_name.lower().endswith(".xlsx"):
        await update.message.reply_text(
            "❌ Пришли именно Excel-файл .xlsx",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    temp_path = "warehouse_upload_temp.xlsx"

    try:
        telegram_file = (
            await context.bot.get_file(
                document.file_id
            )
        )

        await telegram_file.download_to_drive(
            temp_path
        )

        # Проверяем загруженный файл.
        test_dataframe = pd.read_excel(
            temp_path
        )

        test_dataframe.columns = [
            str(column).strip()
            for column in test_dataframe.columns
        ]

        missing_columns = [
            column
            for column in REQUIRED_COLUMNS
            if column not in test_dataframe.columns
        ]

        if missing_columns:
            if os.path.exists(temp_path):
                os.remove(temp_path)

            await update.message.reply_text(
                "⚠️ В Excel не хватает "
                "обязательных колонок:\n"
                + ", ".join(missing_columns),
                reply_markup=MAIN_KEYBOARD,
            )
            return

        # Price, PhotoID, SoldTo, SoldDate и Notes
        # здесь не проверяются, потому что они необязательные.

        async with excel_lock:
            if os.path.exists(FILE_PATH):
                create_backup(
                    "before_upload"
                )

            shutil.move(
                temp_path,
                FILE_PATH,
            )

            # Проверяем, что новый файл нормально читается.
            load_df()

    except Exception as error:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass

        await update.message.reply_text(
            "⚠️ Не удалось загрузить Excel:\n"
            f"{error}",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    context.user_data["mode"] = None

    await update.message.reply_text(
        "✅ Таблица обновлена!\n"
        "Теперь можно искать и удалять позиции.",
        reply_markup=MAIN_KEYBOARD,
    )


# =========================================================
# ПОЛУЧЕНИЕ PHOTO ID
# =========================================================

async def handle_photo(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    if not update.message.photo:
        return

    photo = update.message.photo[-1]

    await update.message.reply_text(
        f"PhotoID:\n{photo.file_id}\n\n"
        "Скопируй этот PhotoID и вставь "
        "в колонку PhotoID в Excel.",
        reply_markup=MAIN_KEYBOARD,
    )


# =========================================================
# ПОИСК ЗАПЧАСТИ
# =========================================================

async def search_part(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    text: str,
):
    query_normalized = (
        normalize_part_for_search(text)
    )

    if not query_normalized:
        await update.message.reply_text(
            "❓ Напиши номер детали.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    try:
        dataframe = load_df()

    except Exception as error:
        await update.message.reply_text(
            f"⚠️ Ошибка: {error}",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    # Точное совпадение.
    exact_matches = dataframe[
        dataframe["_pn_norm"]
        == query_normalized
    ]

    if not exact_matches.empty:
        if len(exact_matches) == 1:
            await send_part_response(
                update,
                context,
                exact_matches.iloc[0],
            )
            return

        responses = [
            fmt_row(row)
            for _, row
            in exact_matches.head(10).iterrows()
        ]

        message = "\n\n".join(responses)

        if len(exact_matches) > 10:
            message += (
                "\n\nℹ️ Найдено больше 10 позиций. "
                "Показаны первые 10."
            )

        await update.message.reply_text(
            message,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    # Частичное совпадение.
    partial_matches = dataframe[
        dataframe["_pn_norm"].str.contains(
            query_normalized,
            na=False,
            regex=False,
        )
    ]

    if not partial_matches.empty:
        if len(partial_matches) == 1:
            await send_part_response(
                update,
                context,
                partial_matches.iloc[0],
            )
            return

        responses = [
            fmt_row(row)
            for _, row
            in partial_matches.head(10).iterrows()
        ]

        message = "\n\n".join(responses)

        if len(partial_matches) > 10:
            message += (
                "\n\nℹ️ Найдено больше 10 вариантов. "
                "Показаны первые 10."
            )

        await update.message.reply_text(
            message,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    # Нечёткий поиск похожих PartNumber.
    part_numbers = (
        dataframe["_pn_norm"]
        .dropna()
        .astype(str)
        .tolist()
    )

    close_matches = difflib.get_close_matches(
        query_normalized,
        part_numbers,
        n=10,
        cutoff=0.75,
    )

    if close_matches:
        fuzzy_matches = dataframe[
            dataframe["_pn_norm"].isin(
                close_matches
            )
        ]

        responses = [
            fmt_row(row)
            for _, row
            in fuzzy_matches.head(10).iterrows()
        ]

        message = (
            "🤔 Точного совпадения нет, "
            "но найдены похожие:\n\n"
            + "\n\n".join(responses)
        )

        await update.message.reply_text(
            message,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    await update.message.reply_text(
        "❓ Ничего не найдено "
        "по этому запросу.",
        reply_markup=MAIN_KEYBOARD,
    )


# =========================================================
# ОБРАБОТКА КНОПОК И СООБЩЕНИЙ
# =========================================================

async def handle_message(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    text = safe_str(
        update.message.text
    )

    if not text:
        return

    if text == "🔍 Найти запчасть":
        await begin_search_mode(
            update,
            context,
        )
        return

    if text == "🗑 Удалить запчасть":
        await begin_delete_mode(
            update,
            context,
        )
        return

    if text == "📥 Скачать Excel":
        context.user_data["mode"] = None

        await download_excel(
            update,
            context,
        )
        return

    if text == "↩️ Отменить удаление":
        await undo_cmd(
            update,
            context,
        )
        return

    if text == "❌ Отмена":
        await cancel_mode(
            update,
            context,
        )
        return

    current_mode = context.user_data.get(
        "mode"
    )

    if current_mode == "delete":
        await prepare_delete(
            update,
            context,
            text,
        )
        return

    if current_mode == "search":
        await search_part(
            update,
            context,
            text,
        )
        return

    # Даже без выбора режима обычное сообщение
    # используется как поиск PartNumber.
    await search_part(
        update,
        context,
        text,
    )


# =========================================================
# ЗАПУСК
# =========================================================

def main():
    if not TOKEN:
        raise RuntimeError(
            "TOKEN не задан. "
            "Добавь TOKEN в Railway Variables."
        )

    application = (
        ApplicationBuilder()
        .token(TOKEN)
        .build()
    )

    application.add_handler(
        CommandHandler(
            "start",
            start,
        )
    )

    application.add_handler(
        CommandHandler(
            "help",
            help_cmd,
        )
    )

    application.add_handler(
        CommandHandler(
            "delete",
            delete_cmd,
        )
    )

    application.add_handler(
        CommandHandler(
            "undo",
            undo_cmd,
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            delete_callback,
            pattern=(
                r"^(confirm_delete|cancel_delete)$"
            ),
        )
    )

    application.add_handler(
        MessageHandler(
            filters.Document.ALL,
            handle_document,
        )
    )

    application.add_handler(
        MessageHandler(
            filters.PHOTO,
            handle_photo,
        )
    )

    application.add_handler(
        MessageHandler(
            filters.TEXT
            & ~filters.COMMAND,
            handle_message,
        )
    )

    print("🤖 Warehouse bot started")

    application.run_polling(
        drop_pending_updates=True
    )


if __name__ == "__main__":
    main()    "Quantity",
    "Shelf",
    "Location",
    "Passport",
    "Category",
    "SerialNumber",
    "Check",
    "Price",
    "PhotoID",
    "SoldTo",
    "SoldDate",
    "Notes",


excel_lock = asyncio.Lock()


# =========================================================
# КНОПКИ ГЛАВНОГО МЕНЮ
# =========================================================

MAIN_KEYBOARD = ReplyKeyboardMarkup(
    [
        [
            "🔍 Найти запчасть",
            "🗑 Удалить запчасть",
        ],
        [
            "📥 Скачать Excel",
            "↩️ Отменить удаление",
        ],
        [
            "❌ Отмена",
        ],
    ],
    resize_keyboard=True,
    is_persistent=True,
)


# =========================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =========================================================

def is_authorized(user_id: int) -> bool:
    """
    Если ADMIN_ID не указан, доступ открыт всем.
    Если указан — удалять и скачивать Excel может только администратор.
    """

    if not ADMIN_ID:
        return True

    return str(user_id) == str(ADMIN_ID)


def normalize_part_for_search(value) -> str:
    """
    Нормализация PartNumber.

    ПУ-11, ПУ 11 и ПУ_11 будут считаться одинаковыми.
    """

    if value is None:
        return ""

    value = str(value).strip().upper()

    value = re.sub(
        r"[\s\-_./\\]+",
        "",
        value,
    )

    return value


def safe_str(value) -> str:
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    return str(value).strip()


def translate_value(value, field):
    value_text = safe_str(value).lower()

    if field == "passport":
        if value_text in ["yes", "y", "true", "1"]:
            return "есть"

        if value_text in ["no", "n", "false", "0"]:
            return "нет"

    if field == "check":
        if value_text in ["yes", "y", "true", "1"]:
            return "проверена"

        if value_text in ["no", "n", "false", "0"]:
            return "не проверена"

    if field == "category":
        if value_text == "new":
            return "новая"

        if value_text == "used":
            return "б/у"

        if value_text == "serviceable":
            return "исправная"

        if value_text == "overhauled":
            return "после ремонта"

    return safe_str(value)


def clean_serial(value) -> str:
    serial = safe_str(value)

    if not serial or serial in ["/", "-", "—"]:
        return "—"

    return serial


def clean_price(value) -> str:
    price = safe_str(value)

    if not price or price in ["/", "-", "—"]:
        return "—"

    price = price.replace("USD", "$")
    price = price.replace("usd", "$")

    return price.strip()


def clean_date(value) -> str:
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    try:
        date_value = pd.to_datetime(value)

        return date_value.strftime(
            "%Y-%m-%d"
        )

    except Exception:
        return str(value).split(" ")[0]


def qty_to_number(value) -> float:
    quantity = safe_str(value).replace(",", ".")

    try:
        return float(quantity)

    except Exception:
        return 0.0


# =========================================================
# ЧТЕНИЕ EXCEL
# =========================================================

def load_df() -> pd.DataFrame:
    if not os.path.exists(FILE_PATH):
        raise FileNotFoundError(
            f"Файл {FILE_PATH} не найден. "
            "Отправь Excel .xlsx файлом в бота."
        )

    dataframe = pd.read_excel(FILE_PATH)

    dataframe.columns = [
        str(column).strip()
        for column in dataframe.columns
    ]

    missing_columns = [
        column
        for column in REQUIRED_COLUMNS
        if column not in dataframe.columns
    ]

    if missing_columns:
        raise ValueError(
            "В Excel не хватает колонок:\n"
            + ", ".join(missing_columns)
        )

    dataframe["PartNumber"] = (
        dataframe["PartNumber"]
        .fillna("")
        .astype(str)
    )

    dataframe["_pn_norm"] = dataframe[
        "PartNumber"
    ].apply(
        normalize_part_for_search
    )

    return dataframe


def get_excel_headers(workbook) -> dict:
    worksheet = workbook.active
    headers = {}

    for column_number, cell in enumerate(
        worksheet[1],
        start=1,
    ):
        if cell.value is not None:
            header_name = str(cell.value).strip()
            headers[header_name] = column_number

    return headers


# =========================================================
# РЕЗЕРВНЫЕ КОПИИ
# =========================================================

def create_backup(prefix: str = "backup") -> str:
    if not os.path.exists(FILE_PATH):
        raise FileNotFoundError(
            f"Файл {FILE_PATH} не найден."
        )

    timestamp = datetime.now().strftime(
        "%Y-%m-%d_%H-%M-%S_%f"
    )

    backup_path = (
        f"warehouse_{prefix}_{timestamp}.xlsx"
    )

    shutil.copy2(
        FILE_PATH,
        backup_path,
    )

    return backup_path


# =========================================================
# УДАЛЕНИЕ ПО PARTNUMBER
# =========================================================

def parse_part_numbers(text: str) -> list[str]:
    """
    Можно писать:

    ПУ-11

    Или:

    ПУ-11, ЭЦН-333М, БПСР4

    Или каждый номер с новой строки.
    """

    text = safe_str(text)

    text = re.sub(
        r"^/delete(?:@\w+)?",
        "",
        text,
        flags=re.IGNORECASE,
    ).strip()

    if not text:
        return []

    parts = re.split(
        r"[,;\n]+",
        text,
    )

    result = []
    seen = set()

    for part in parts:
        part = part.strip()

        if not part:
            continue

        normalized = normalize_part_for_search(
            part
        )

        if normalized and normalized not in seen:
            seen.add(normalized)
            result.append(part)

    return result


def find_rows_by_part_numbers(
    requested_parts: list[str],
) -> tuple[list[dict], list[str]]:
    """
    Ищет только точное совпадение PartNumber.

    Если один PartNumber встречается в нескольких строках,
    бот найдёт все такие строки.
    """

    if not os.path.exists(FILE_PATH):
        raise FileNotFoundError(
            f"Файл {FILE_PATH} не найден."
        )

    workbook = load_workbook(
        FILE_PATH,
        data_only=False,
    )

    worksheet = workbook.active
    headers = get_excel_headers(workbook)

    part_column = headers.get("PartNumber")
    serial_column = headers.get("SerialNumber")
    quantity_column = headers.get("Quantity")
    shelf_column = headers.get("Shelf")
    location_column = headers.get("Location")

    if not part_column:
        workbook.close()

        raise ValueError(
            "В Excel не найдена колонка PartNumber."
        )

    requested_map = {
        normalize_part_for_search(part): part
        for part in requested_parts
    }

    found_normalized = set()
    found_rows = []

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        part_value = worksheet.cell(
            row=row_number,
            column=part_column,
        ).value

        normalized = normalize_part_for_search(
            part_value
        )

        if normalized not in requested_map:
            continue

        found_normalized.add(normalized)

        serial_value = ""

        if serial_column:
            serial_value = worksheet.cell(
                row=row_number,
                column=serial_column,
            ).value

        quantity_value = ""

        if quantity_column:
            quantity_value = worksheet.cell(
                row=row_number,
                column=quantity_column,
            ).value

        shelf_value = ""

        if shelf_column:
            shelf_value = worksheet.cell(
                row=row_number,
                column=shelf_column,
            ).value

        location_value = ""

        if location_column:
            location_value = worksheet.cell(
                row=row_number,
                column=location_column,
            ).value

        found_rows.append(
            {
                "row": row_number,
                "part_number": safe_str(part_value),
                "normalized": normalized,
                "serial_number": safe_str(
                    serial_value
                ),
                "quantity": safe_str(
                    quantity_value
                ),
                "shelf": safe_str(
                    shelf_value
                ),
                "location": safe_str(
                    location_value
                ),
            }
        )

    workbook.close()

    not_found = []

    for normalized, original in requested_map.items():
        if normalized not in found_normalized:
            not_found.append(original)

    return found_rows, not_found


def verify_pending_rows(
    pending_rows: list[dict],
) -> bool:
    """
    Проверяет, не изменился ли Excel между поиском
    и нажатием кнопки «Удалить».
    """

    if not os.path.exists(FILE_PATH):
        return False

    workbook = load_workbook(
        FILE_PATH,
        data_only=False,
    )

    worksheet = workbook.active
    headers = get_excel_headers(workbook)

    part_column = headers.get("PartNumber")

    if not part_column:
        workbook.close()
        return False

    for item in pending_rows:
        row_number = item["row"]

        if (
            row_number < 2
            or row_number > worksheet.max_row
        ):
            workbook.close()
            return False

        current_part = worksheet.cell(
            row=row_number,
            column=part_column,
        ).value

        current_normalized = (
            normalize_part_for_search(
                current_part
            )
        )

        if (
            current_normalized
            != item["normalized"]
        ):
            workbook.close()
            return False

    workbook.close()

    return True


def delete_excel_rows(
    rows_to_delete: list[dict],
) -> list[dict]:
    """
    Полностью удаляет строки из Excel.

    Удаление выполняется снизу вверх,
    чтобы номера строк не смещались.
    """

    if not os.path.exists(FILE_PATH):
        raise FileNotFoundError(
            f"Файл {FILE_PATH} не найден."
        )

    workbook = load_workbook(FILE_PATH)
    worksheet = workbook.active

    deleted_items = []

    sorted_rows = sorted(
        rows_to_delete,
        key=lambda item: item["row"],
        reverse=True,
    )

    for item in sorted_rows:
        row_number = item["row"]

        if (
            row_number < 2
            or row_number > worksheet.max_row
        ):
            continue

        deleted_items.append(item)

        worksheet.delete_rows(
            row_number,
            1,
        )

    workbook.save(FILE_PATH)
    workbook.close()

    return sorted(
        deleted_items,
        key=lambda item: item["row"],
    )


def format_delete_preview(
    found_rows: list[dict],
    not_found: list[str],
) -> str:
    lines = [
        "⚠️ Будут полностью удалены:",
        "",
    ]

    for item in found_rows[:40]:
        line = f'📦 {item["part_number"]}'

        if item["serial_number"]:
            line += (
                f' | S/N: {item["serial_number"]}'
            )

        if item["quantity"]:
            line += (
                f' | Qty: {item["quantity"]}'
            )

        if item["shelf"]:
            line += (
                f' | Полка: {item["shelf"]}'
            )

        if item["location"]:
            line += (
                f' | Ячейка: {item["location"]}'
            )

        lines.append(line)

    if len(found_rows) > 40:
        lines.append("")
        lines.append(
            f"…и ещё {len(found_rows) - 40} строк."
        )

    if not_found:
        lines.append("")
        lines.append("❌ Не найдены:")

        for part in not_found[:30]:
            lines.append(f"• {part}")

        if len(not_found) > 30:
            lines.append(
                f"…и ещё {len(not_found) - 30}."
            )

    lines.append("")
    lines.append(
        f"Всего будет удалено строк: "
        f"{len(found_rows)}"
    )

    lines.append("")
    lines.append(
        "После удаления эти запчасти исчезнут "
        "из Excel и из поиска бота."
    )

    return "\n".join(lines)


# =========================================================
# ФОРМАТ ОТВЕТА ПРИ ПОИСКЕ
# =========================================================

def fmt_row(row) -> str:
    part = safe_str(
        row.get("PartNumber")
    )

    quantity = safe_str(
        row.get("Quantity")
    )

    shelf = safe_str(
        row.get("Shelf")
    )

    location = safe_str(
        row.get("Location")
    )

    passport = translate_value(
        row.get("Passport"),
        "passport",
    )

    category = translate_value(
        row.get("Category"),
        "category",
    )

    check = translate_value(
        row.get("Check"),
        "check",
    )

    serial = clean_serial(
        row.get("SerialNumber")
    )

    price = clean_price(
        row.get("Price")
    )

    sold_to = safe_str(
        row.get("SoldTo")
    )

    sold_date = clean_date(
        row.get("SoldDate")
    )

    notes = safe_str(
        row.get("Notes")
    )

    quantity_number = qty_to_number(
        row.get("Quantity")
    )

    if quantity_number <= 0:
        text = (
            f"❌ ПРОДАНО\n"
            f"📦 {part}\n"
            f"📍 Полка: {shelf}, "
            f"ячейка: {location}\n"
            f"🔢 Количество: {quantity}\n"
            f"📄 Паспорт: {passport}\n"
            f"🆕 Категория: {category}\n"
            f"💰 Цена: {price}\n"
            f"🔑 Серийный номер: {serial}\n"
            f"✔ Проверка: {check}"
        )

        if sold_to:
            text += (
                f"\n👤 Кому продано: {sold_to}"
            )

        if sold_date:
            text += (
                f"\n📅 Дата продажи: {sold_date}"
            )

        if notes:
            text += (
                f"\n📝 Заметка: {notes}"
            )

        return text

    return (
        f"✅ {part} есть в наличии\n"
        f"📦 Полка: {shelf}, "
        f"ячейка: {location}\n"
        f"🔢 Количество: {quantity}\n"
        f"📄 Паспорт: {passport}\n"
        f"🆕 Категория: {category}\n"
        f"💰 Цена: {price}\n"
        f"🔑 Серийный номер: {serial}\n"
        f"✔ Проверка: {check}"
    )


async def send_part_response(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    row,
):
    caption = fmt_row(row)

    photo_id = safe_str(
        row.get("PhotoID")
    )

    if (
        photo_id
        and photo_id.lower() != "nan"
    ):
        try:
            await update.message.reply_photo(
                photo=photo_id,
                caption=caption,
                reply_markup=MAIN_KEYBOARD,
            )
            return

        except Exception as error:
            print(
                "PHOTO ERROR:",
                error,
            )

    await update.message.reply_text(
        caption,
        reply_markup=MAIN_KEYBOARD,
    )


# =========================================================
# ОТПРАВКА АКТУАЛЬНОГО EXCEL
# =========================================================

async def send_current_excel_to_chat(
    context: ContextTypes.DEFAULT_TYPE,
    chat_id: int,
    caption: str,
):
    if not os.path.exists(FILE_PATH):
        await context.bot.send_message(
            chat_id=chat_id,
            text=(
                "❌ Файл warehouse.xlsx "
                "не найден."
            ),
        )
        return

    with open(FILE_PATH, "rb") as excel_file:
        await context.bot.send_document(
            chat_id=chat_id,
            document=excel_file,
            filename="warehouse_actual.xlsx",
            caption=caption,
            reply_markup=MAIN_KEYBOARD,
        )


async def download_excel(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    user = update.effective_user

    if (
        not user
        or not is_authorized(user.id)
    ):
        await update.message.reply_text(
            "⛔ У вас нет доступа "
            "к скачиванию базы.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    try:
        async with excel_lock:
            await send_current_excel_to_chat(
                context=context,
                chat_id=update.effective_chat.id,
                caption=(
                    "📥 Актуальный склад.\n\n"
                    "В этом файле уже нет "
                    "удалённых запчастей."
                ),
            )

    except Exception as error:
        await update.message.reply_text(
            "⚠️ Не удалось отправить Excel:\n"
            f"{error}",
            reply_markup=MAIN_KEYBOARD,
        )


# =========================================================
# КОМАНДЫ И КНОПКИ
# =========================================================

async def start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    context.user_data["mode"] = None

    context.user_data.pop(
        "pending_delete_rows",
        None,
    )

    await update.message.reply_text(
        "Привет! 👋\n\n"
        "Выбери нужное действие кнопкой внизу.",
        reply_markup=MAIN_KEYBOARD,
    )


async def help_cmd(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    await update.message.reply_text(
        "🔍 Найти запчасть — поиск по PartNumber.\n\n"
        "🗑 Удалить запчасть — удаление "
        "одной или нескольких позиций.\n\n"
        "📥 Скачать Excel — получить "
        "актуальный склад без удалённых позиций.\n\n"
        "↩️ Отменить удаление — восстановить "
        "последнее удаление.\n\n"
        "Чтобы обновить базу, просто отправь "
        "боту новый Excel-файл .xlsx.",
        reply_markup=MAIN_KEYBOARD,
    )


async def begin_search_mode(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    context.user_data["mode"] = "search"

    await update.message.reply_text(
        "🔍 Отправь PartNumber "
        "или часть номера.\n\n"
        "Например:\n"
        "ПУ-11",
        reply_markup=MAIN_KEYBOARD,
    )


async def begin_delete_mode(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    user = update.effective_user

    if (
        not user
        or not is_authorized(user.id)
    ):
        await update.message.reply_text(
            "⛔ У вас нет доступа к удалению.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    context.user_data["mode"] = "delete"

    await update.message.reply_text(
        "🗑 Отправь PartNumber, "
        "который нужно удалить.\n\n"
        "Одна позиция:\n"
        "ПУ-11\n\n"
        "Несколько позиций:\n"
        "ПУ-11, ЭЦН-333М, БПСР4",
        reply_markup=MAIN_KEYBOARD,
    )


async def cancel_mode(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    context.user_data["mode"] = None

    context.user_data.pop(
        "pending_delete_rows",
        None,
    )

    await update.message.reply_text(
        "❌ Действие отменено.",
        reply_markup=MAIN_KEYBOARD,
    )


# =========================================================
# ПОДГОТОВКА УДАЛЕНИЯ
# =========================================================

async def prepare_delete(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    text: str,
):
    requested_parts = parse_part_numbers(
        text
    )

    if not requested_parts:
        await update.message.reply_text(
            "❌ Не удалось прочитать PartNumber.\n\n"
            "Например:\n"
            "ПУ-11, ЭЦН-333М",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    if len(requested_parts) > 500:
        await update.message.reply_text(
            "❌ За один раз можно указать "
            "максимум 500 PartNumber.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    try:
        async with excel_lock:
            found_rows, not_found = (
                find_rows_by_part_numbers(
                    requested_parts
                )
            )

    except Exception as error:
        await update.message.reply_text(
            "⚠️ Ошибка при чтении Excel:\n"
            f"{error}",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    if not found_rows:
        message = (
            "❌ Ни одна запчасть не найдена.\n\n"
            "Проверь написание PartNumber."
        )

        if not_found:
            message += "\n\nНе найдены:\n"

            message += "\n".join(
                f"• {part}"
                for part in not_found[:30]
            )

        await update.message.reply_text(
            message,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    context.user_data[
        "pending_delete_rows"
    ] = found_rows

    keyboard = InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "✅ Удалить",
                    callback_data="confirm_delete",
                ),
                InlineKeyboardButton(
                    "❌ Отмена",
                    callback_data="cancel_delete",
                ),
            ]
        ]
    )

    await update.message.reply_text(
        format_delete_preview(
            found_rows,
            not_found,
        ),
        reply_markup=keyboard,
    )


async def delete_cmd(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    user = update.effective_user

    if (
        not user
        or not is_authorized(user.id)
    ):
        await update.message.reply_text(
            "⛔ У вас нет доступа к удалению.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    command_text = update.message.text or ""

    part_text = re.sub(
        r"^/delete(?:@\w+)?",
        "",
        command_text,
        flags=re.IGNORECASE,
    ).strip()

    if not part_text:
        await begin_delete_mode(
            update,
            context,
        )
        return

    await prepare_delete(
        update,
        context,
        part_text,
    )


# =========================================================
# ПОДТВЕРЖДЕНИЕ УДАЛЕНИЯ
# =========================================================

async def delete_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    query = update.callback_query

    if not query:
        return

    await query.answer()

    user = query.from_user

    if not is_authorized(user.id):
        await query.edit_message_text(
            "⛔ У вас нет доступа к удалению."
        )
        return

    if query.data == "cancel_delete":
        context.user_data["mode"] = None

        context.user_data.pop(
            "pending_delete_rows",
            None,
        )

        await query.edit_message_text(
            "❌ Удаление отменено. "
            "Excel не изменён."
        )
        return

    if query.data != "confirm_delete":
        return

    pending_rows = context.user_data.get(
        "pending_delete_rows"
    )

    if not pending_rows:
        await query.edit_message_text(
            "⚠️ Список удаления уже недействителен.\n"
            "Нажми кнопку «🗑 Удалить запчасть» "
            "ещё раз."
        )
        return

    try:
        async with excel_lock:
            if not verify_pending_rows(
                pending_rows
            ):
                context.user_data.pop(
                    "pending_delete_rows",
                    None,
                )

                context.user_data["mode"] = None

                await query.edit_message_text(
                    "⚠️ Excel изменился после поиска.\n"
                    "Для безопасности удаление "
                    "остановлено.\n\n"
                    "Попробуй ещё раз."
                )
                return

            backup_path = create_backup(
                "before_delete"
            )

            deleted_items = delete_excel_rows(
                pending_rows
            )

            context.application.bot_data[
                "last_backup"
            ] = backup_path

            context.application.bot_data[
                "last_delete_user"
            ] = user.id

    except Exception as error:
        await query.edit_message_text(
            "⚠️ Ошибка при удалении:\n"
            f"{error}"
        )
        return

    context.user_data.pop(
        "pending_delete_rows",
        None,
    )

    context.user_data["mode"] = None

    lines = [
        "✅ Полностью удалено строк: "
        f"{len(deleted_items)}",
        "",
    ]

    for item in deleted_items[:40]:
        line = f'• {item["part_number"]}'

        if item["serial_number"]:
            line += (
                f' | S/N: {item["serial_number"]}'
            )

        lines.append(line)

    if len(deleted_items) > 40:
        lines.append("")
        lines.append(
            f"…и ещё {len(deleted_items) - 40}."
        )

    lines.append("")
    lines.append(
        "Эти запчасти больше не находятся "
        "через поиск."
    )

    lines.append("")
    lines.append(
        "Сейчас отправляю новый актуальный Excel."
    )

    await query.edit_message_text(
        "\n".join(lines)
    )

    try:
        await send_current_excel_to_chat(
            context=context,
            chat_id=query.message.chat_id,
            caption=(
                "📥 Обновлённый Excel.\n\n"
                "Удалённые запчасти уже убраны "
                "из списка.\n"
                "Этот файл можно отправлять клиентам."
            ),
        )

    except Exception as error:
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=(
                "⚠️ Запчасти удалены, но не удалось "
                "отправить Excel:\n"
                f"{error}\n\n"
                "Нажми кнопку «📥 Скачать Excel»."
            ),
            reply_markup=MAIN_KEYBOARD,
        )


# =========================================================
# ОТМЕНА ПОСЛЕДНЕГО УДАЛЕНИЯ
# =========================================================

async def undo_cmd(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    user = update.effective_user

    if (
        not user
        or not is_authorized(user.id)
    ):
        await update.message.reply_text(
            "⛔ У вас нет доступа "
            "к восстановлению.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    backup_path = (
        context.application.bot_data.get(
            "last_backup"
        )
    )

    if (
        not backup_path
        or not os.path.exists(backup_path)
    ):
        await update.message.reply_text(
            "❌ Нет последнего удаления, "
            "которое можно отменить.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    try:
        async with excel_lock:
            create_backup(
                "before_undo"
            )

            shutil.copy2(
                backup_path,
                FILE_PATH,
            )

            context.application.bot_data.pop(
                "last_backup",
                None,
            )

            context.application.bot_data.pop(
                "last_delete_user",
                None,
            )

    except Exception as error:
        await update.message.reply_text(
            "⚠️ Не удалось восстановить Excel:\n"
            f"{error}",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    context.user_data["mode"] = None

    await update.message.reply_text(
        "✅ Последнее удаление отменено.\n"
        "Предыдущая версия Excel восстановлена.",
        reply_markup=MAIN_KEYBOARD,
    )

    try:
        await send_current_excel_to_chat(
            context=context,
            chat_id=update.effective_chat.id,
            caption=(
                "📥 Восстановленный Excel.\n\n"
                "Удалённые позиции снова находятся "
                "в списке."
            ),
        )

    except Exception as error:
        await update.message.reply_text(
            "⚠️ Excel восстановлен, "
            "но не удалось отправить файл:\n"
            f"{error}",
            reply_markup=MAIN_KEYBOARD,
        )


# =========================================================
# ЗАГРУЗКА НОВОГО EXCEL В БОТА
# =========================================================

async def handle_document(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    user = update.effective_user

    if (
        not user
        or not is_authorized(user.id)
    ):
        await update.message.reply_text(
            "⛔ У вас нет доступа "
            "к обновлению базы.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    document = update.message.document

    if not document:
        return

    file_name = document.file_name or ""

    if not file_name.lower().endswith(
        ".xlsx"
    ):
        await update.message.reply_text(
            "❌ Пришли именно Excel-файл .xlsx",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    temp_path = (
        "warehouse_upload_temp.xlsx"
    )

    try:
        telegram_file = (
            await context.bot.get_file(
                document.file_id
            )
        )

        await telegram_file.download_to_drive(
            temp_path
        )

        test_dataframe = pd.read_excel(
            temp_path
        )

        test_dataframe.columns = [
            str(column).strip()
            for column in test_dataframe.columns
        ]

        missing_columns = [
            column
            for column in REQUIRED_COLUMNS
            if column
            not in test_dataframe.columns
        ]

        if missing_columns:
            if os.path.exists(temp_path):
                os.remove(temp_path)

            await update.message.reply_text(
                "⚠️ В Excel не хватает колонок:\n"
                + ", ".join(missing_columns),
                reply_markup=MAIN_KEYBOARD,
            )
            return

        async with excel_lock:
            if os.path.exists(FILE_PATH):
                create_backup(
                    "before_upload"
                )

            shutil.move(
                temp_path,
                FILE_PATH,
            )

            load_df()

    except Exception as error:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass

        await update.message.reply_text(
            "⚠️ Не удалось загрузить Excel:\n"
            f"{error}",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    context.user_data["mode"] = None

    await update.message.reply_text(
        "✅ Таблица обновлена!\n"
        "Теперь можно искать и удалять позиции.",
        reply_markup=MAIN_KEYBOARD,
    )


# =========================================================
# ПОЛУЧЕНИЕ PHOTO ID
# =========================================================

async def handle_photo(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    if not update.message.photo:
        return

    photo = update.message.photo[-1]

    await update.message.reply_text(
        f"PhotoID:\n{photo.file_id}\n\n"
        "Скопируй этот PhotoID и вставь "
        "в колонку PhotoID в Excel.",
        reply_markup=MAIN_KEYBOARD,
    )


# =========================================================
# ПОИСК ЗАПЧАСТИ
# =========================================================

async def search_part(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    text: str,
):
    query_normalized = (
        normalize_part_for_search(text)
    )

    if not query_normalized:
        await update.message.reply_text(
            "❓ Напиши номер детали.",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    try:
        dataframe = load_df()

    except Exception as error:
        await update.message.reply_text(
            f"⚠️ Ошибка: {error}",
            reply_markup=MAIN_KEYBOARD,
        )
        return

    # Точное совпадение.
    exact_matches = dataframe[
        dataframe["_pn_norm"]
        == query_normalized
    ]

    if not exact_matches.empty:
        if len(exact_matches) == 1:
            await send_part_response(
                update,
                context,
                exact_matches.iloc[0],
            )
            return

        responses = [
            fmt_row(row)
            for _, row
            in exact_matches.head(10).iterrows()
        ]

        message = "\n\n".join(
            responses
        )

        if len(exact_matches) > 10:
            message += (
                "\n\nℹ️ Найдено больше 10 позиций. "
                "Показаны первые 10."
            )

        await update.message.reply_text(
            message,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    # Частичное совпадение.
    partial_matches = dataframe[
        dataframe["_pn_norm"].str.contains(
            query_normalized,
            na=False,
            regex=False,
        )
    ]

    if not partial_matches.empty:
        if len(partial_matches) == 1:
            await send_part_response(
                update,
                context,
                partial_matches.iloc[0],
            )
            return

        responses = [
            fmt_row(row)
            for _, row
            in partial_matches.head(10).iterrows()
        ]

        message = "\n\n".join(
            responses
        )

        if len(partial_matches) > 10:
            message += (
                "\n\nℹ️ Найдено больше 10 вариантов. "
                "Показаны первые 10."
            )

        await update.message.reply_text(
            message,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    # Поиск похожих номеров.
    part_numbers = (
        dataframe["_pn_norm"]
        .dropna()
        .astype(str)
        .tolist()
    )

    close_matches = difflib.get_close_matches(
        query_normalized,
        part_numbers,
        n=10,
        cutoff=0.75,
    )

    if close_matches:
        fuzzy_matches = dataframe[
            dataframe["_pn_norm"].isin(
                close_matches
            )
        ]

        responses = [
            fmt_row(row)
            for _, row
            in fuzzy_matches.head(10).iterrows()
        ]

        message = (
            "🤔 Точного совпадения нет, "
            "но найдены похожие:\n\n"
            + "\n\n".join(responses)
        )

        await update.message.reply_text(
            message,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    await update.message.reply_text(
        "❓ Ничего не найдено "
        "по этому запросу.",
        reply_markup=MAIN_KEYBOARD,
    )


# =========================================================
# ОБРАБОТКА ОБЫЧНЫХ СООБЩЕНИЙ И КНОПОК
# =========================================================

async def handle_message(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
):
    text = safe_str(
        update.message.text
    )

    if not text:
        return

    if text == "🔍 Найти запчасть":
        await begin_search_mode(
            update,
            context,
        )
        return

    if text == "🗑 Удалить запчасть":
        await begin_delete_mode(
            update,
            context,
        )
        return

    if text == "📥 Скачать Excel":
        context.user_data["mode"] = None

        await download_excel(
            update,
            context,
        )
        return

    if text == "↩️ Отменить удаление":
        await undo_cmd(
            update,
            context,
        )
        return

    if text == "❌ Отмена":
        await cancel_mode(
            update,
            context,
        )
        return

    current_mode = context.user_data.get(
        "mode"
    )

    if current_mode == "delete":
        await prepare_delete(
            update,
            context,
            text,
        )
        return

    if current_mode == "search":
        await search_part(
            update,
            context,
            text,
        )
        return

    # Даже без нажатия кнопки обычный текст ищет запчасть.
    await search_part(
        update,
        context,
        text,
    )


# =========================================================
# ЗАПУСК БОТА
# =========================================================

def main():
    if not TOKEN:
        raise RuntimeError(
            "TOKEN не задан.\n"
            "Добавь TOKEN в Railway Variables."
        )

    application = (
        ApplicationBuilder()
        .token(TOKEN)
        .build()
    )

    application.add_handler(
        CommandHandler(
            "start",
            start,
        )
    )

    application.add_handler(
        CommandHandler(
            "help",
            help_cmd,
        )
    )

    application.add_handler(
        CommandHandler(
            "delete",
            delete_cmd,
        )
    )

    application.add_handler(
        CommandHandler(
            "undo",
            undo_cmd,
        )
    )

    application.add_handler(
        CallbackQueryHandler(
            delete_callback,
            pattern=(
                r"^(confirm_delete|cancel_delete)$"
            ),
        )
    )

    application.add_handler(
        MessageHandler(
            filters.Document.ALL,
            handle_document,
        )
    )

    application.add_handler(
        MessageHandler(
            filters.PHOTO,
            handle_photo,
        )
    )

    application.add_handler(
        MessageHandler(
            filters.TEXT
            & ~filters.COMMAND,
            handle_message,
        )
    )

    print(
        "🤖 Warehouse bot started"
    )

    application.run_polling(
        drop_pending_updates=True
    )


if __name__ == "__main__":
    main()
